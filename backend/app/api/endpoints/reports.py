# ============================================================
# PetroVision — Advanced Reporting & Export Endpoints
# ============================================================
# Shift / daily / weekly report generation.
# KPI summaries, alarm statistics, batch analytics.
# Export to JSON preview + Excel (.xlsx) download.
# ============================================================

import io
import math
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case, and_
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.batch import Batch, ProductionReport
from app.models.alarm import Alarm
from app.models.instrument import Instrument

router = APIRouter(prefix="/reports", tags=["Reportes Avanzados"])


# ── Helpers ──────────────────────────────────────────────────

PERIOD_MAP = {
    "shift": timedelta(hours=8),
    "daily": timedelta(days=1),
    "weekly": timedelta(weeks=1),
    "monthly": timedelta(days=30),
}


def _compute_kpis(db: Session, process_id: int, start: datetime, end: datetime) -> dict:
    """Compute KPIs for a given period."""
    batches = db.query(Batch).filter(
        Batch.process_id == process_id,
        Batch.created_at >= start,
        Batch.created_at <= end,
    ).all()

    completed = [b for b in batches if b.status in ("COMPLETED", "APPROVED")]

    total_production = sum(b.product_amount_kg or 0 for b in completed)
    total_feed = sum(b.feed_amount_kg or 0 for b in completed)
    yields = [b.yield_actual for b in completed if b.yield_actual is not None]
    purities = [b.purity for b in completed if b.purity is not None]
    costs = [b.production_cost for b in completed if b.production_cost is not None]
    revenues = [b.revenue for b in completed if b.revenue is not None]
    temps = [b.avg_temperature for b in completed if b.avg_temperature is not None]
    pressures = [b.avg_pressure for b in completed if b.avg_pressure is not None]

    avg_yield = sum(yields) / len(yields) if yields else None
    avg_purity = sum(purities) / len(purities) if purities else None
    total_cost = sum(costs)
    total_revenue = sum(revenues)

    # Quality breakdown
    grades = {}
    for b in completed:
        g = b.quality_grade or "N/A"
        grades[g] = grades.get(g, 0) + 1

    # OEE estimate (simplified)
    period_hours = (end - start).total_seconds() / 3600
    operating_hours = 0
    for b in completed:
        if b.actual_start and b.actual_end:
            operating_hours += (b.actual_end - b.actual_start).total_seconds() / 3600
    oee = (operating_hours / period_hours * 100) if period_hours > 0 else None

    return {
        "total_batches": len(batches),
        "completed_batches": len(completed),
        "planned_batches": sum(1 for b in batches if b.status == "PLANNED"),
        "in_progress_batches": sum(1 for b in batches if b.status == "IN_PROGRESS"),
        "total_production_kg": round(total_production, 2),
        "total_feed_kg": round(total_feed, 2),
        "avg_yield": round(avg_yield, 2) if avg_yield else None,
        "avg_purity": round(avg_purity, 2) if avg_purity else None,
        "total_cost": round(total_cost, 2),
        "total_revenue": round(total_revenue, 2),
        "margin": round(total_revenue - total_cost, 2),
        "margin_pct": round((total_revenue - total_cost) / total_revenue * 100, 1) if total_revenue > 0 else None,
        "avg_temperature": round(sum(temps) / len(temps), 1) if temps else None,
        "avg_pressure": round(sum(pressures) / len(pressures), 2) if pressures else None,
        "quality_breakdown": grades,
        "oee": round(oee, 1) if oee is not None else None,
        "operating_hours": round(operating_hours, 1),
    }


def _alarm_stats(db: Session, process_id: int, start: datetime, end: datetime) -> dict:
    """Alarm statistics for a period."""
    alarms = db.query(Alarm).filter(
        Alarm.process_id == process_id,
        Alarm.triggered_at >= start,
        Alarm.triggered_at <= end,
    ).all()

    total = len(alarms)
    by_priority = {}
    by_type = {}
    by_instrument = {}
    response_times = []

    for a in alarms:
        by_priority[a.priority] = by_priority.get(a.priority, 0) + 1
        by_type[a.alarm_type] = by_type.get(a.alarm_type, 0) + 1
        by_instrument[a.instrument_tag] = by_instrument.get(a.instrument_tag, 0) + 1
        if a.acknowledged_at and a.triggered_at:
            rt = (a.acknowledged_at - a.triggered_at).total_seconds()
            response_times.append(rt)

    avg_response = sum(response_times) / len(response_times) if response_times else None

    # Top offenders
    top_instruments = sorted(by_instrument.items(), key=lambda x: -x[1])[:10]

    return {
        "total": total,
        "by_priority": by_priority,
        "by_type": by_type,
        "top_instruments": [{"tag": t, "count": c} for t, c in top_instruments],
        "avg_response_time_s": round(avg_response, 1) if avg_response is not None else None,
        "acknowledged_pct": round(
            sum(1 for a in alarms if a.acknowledged_at) / total * 100, 1
        ) if total > 0 else None,
        "critical": by_priority.get("CRITICA", 0),
        "high": by_priority.get("ALTA", 0),
    }


def _batch_details(db: Session, process_id: int, start: datetime, end: datetime) -> list[dict]:
    """Detailed batch data for a period."""
    batches = db.query(Batch).filter(
        Batch.process_id == process_id,
        Batch.created_at >= start,
        Batch.created_at <= end,
    ).order_by(Batch.created_at).all()

    return [
        {
            "batch_number": b.batch_number,
            "status": b.status,
            "actual_start": b.actual_start.isoformat() if b.actual_start else None,
            "actual_end": b.actual_end.isoformat() if b.actual_end else None,
            "feed_kg": b.feed_amount_kg,
            "product_kg": b.product_amount_kg,
            "yield": b.yield_actual,
            "purity": b.purity,
            "quality_grade": b.quality_grade,
            "avg_temp": b.avg_temperature,
            "avg_press": b.avg_pressure,
            "cost": b.production_cost,
            "revenue": b.revenue,
        }
        for b in batches
    ]


# ── 1. Report Preview (JSON) ────────────────────────────────

@router.get("/generate")
def generate_report(
    process_id: int = Query(1, description="Process ID"),
    report_type: str = Query("daily", description="shift | daily | weekly | monthly"),
    periods_back: int = Query(1, ge=1, le=12, description="Number of periods back"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate a production report preview."""
    delta = PERIOD_MAP.get(report_type, timedelta(days=1))
    now = datetime.now(timezone.utc)
    end = now
    start = now - delta * periods_back

    kpis = _compute_kpis(db, process_id, start, end)
    alarms = _alarm_stats(db, process_id, start, end)
    batches = _batch_details(db, process_id, start, end)

    # Save to DB
    pr = ProductionReport(
        process_id=process_id,
        report_type=report_type,
        period_start=start,
        period_end=end,
        total_production_kg=kpis["total_production_kg"],
        total_feed_kg=kpis["total_feed_kg"],
        avg_yield=kpis["avg_yield"],
        avg_purity=kpis["avg_purity"],
        oee=kpis["oee"],
        uptime_hours=kpis["operating_hours"],
        total_cost=kpis["total_cost"],
        total_revenue=kpis["total_revenue"],
        margin=kpis["margin_pct"],
        total_alarms=alarms["total"],
        critical_alarms=alarms["critical"],
        generated_by=current_user.id,
        data={"kpis": kpis, "alarms": alarms, "batches": batches},
    )
    db.add(pr)
    db.commit()
    db.refresh(pr)

    return {
        "report_id": pr.id,
        "report_type": report_type,
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "kpis": kpis,
        "alarms": alarms,
        "batches": batches,
        "generated_at": pr.generated_at.isoformat() if pr.generated_at else None,
        "generated_by": current_user.username,
    }


# ── 2. Report History ───────────────────────────────────────

@router.get("/history")
def report_history(
    process_id: int = Query(1),
    report_type: Optional[str] = None,
    limit: int = Query(20, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List previously generated reports."""
    q = db.query(ProductionReport).filter(ProductionReport.process_id == process_id)
    if report_type:
        q = q.filter(ProductionReport.report_type == report_type)
    reports = q.order_by(ProductionReport.generated_at.desc()).limit(limit).all()
    return [
        {
            "id": r.id,
            "report_type": r.report_type,
            "period_start": r.period_start.isoformat() if r.period_start else None,
            "period_end": r.period_end.isoformat() if r.period_end else None,
            "total_production_kg": r.total_production_kg,
            "avg_yield": r.avg_yield,
            "avg_purity": r.avg_purity,
            "oee": r.oee,
            "total_alarms": r.total_alarms,
            "critical_alarms": r.critical_alarms,
            "generated_at": r.generated_at.isoformat() if r.generated_at else None,
        }
        for r in reports
    ]


# ── 3. Single Report Detail ─────────────────────────────────

@router.get("/history/{report_id}")
def get_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get a specific report with full data."""
    r = db.query(ProductionReport).filter(ProductionReport.id == report_id).first()
    if not r:
        return {"error": "Reporte no encontrado"}
    return {
        "id": r.id,
        "report_type": r.report_type,
        "period_start": r.period_start.isoformat() if r.period_start else None,
        "period_end": r.period_end.isoformat() if r.period_end else None,
        "kpis": r.data.get("kpis") if r.data else {},
        "alarms": r.data.get("alarms") if r.data else {},
        "batches": r.data.get("batches") if r.data else [],
        "generated_at": r.generated_at.isoformat() if r.generated_at else None,
    }


# ── 4. Export to Excel ───────────────────────────────────────

@router.get("/export/excel")
def export_excel(
    process_id: int = Query(1),
    report_type: str = Query("daily"),
    periods_back: int = Query(1, ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export report data as Excel .xlsx file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    delta = PERIOD_MAP.get(report_type, timedelta(days=1))
    now = datetime.now(timezone.utc)
    end = now
    start = now - delta * periods_back

    kpis = _compute_kpis(db, process_id, start, end)
    alarm_stats = _alarm_stats(db, process_id, start, end)
    batches = _batch_details(db, process_id, start, end)

    wb = openpyxl.Workbook()

    # ── Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="1e3a5f")
    kpi_label_font = Font(name="Calibri", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Sheet 1: KPIs
    ws = wb.active
    ws.title = "KPIs"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"PetroVision — Reporte {report_type.capitalize()}"
    ws["A1"].font = title_font
    ws["A3"] = "Periodo:"
    ws["A3"].font = kpi_label_font
    ws["B3"] = f"{start.strftime('%Y-%m-%d %H:%M')} — {end.strftime('%Y-%m-%d %H:%M')}"

    kpi_rows = [
        ("Producción Total (kg)", kpis["total_production_kg"]),
        ("Alimentación Total (kg)", kpis["total_feed_kg"]),
        ("Rendimiento Promedio (%)", kpis["avg_yield"]),
        ("Pureza Promedio (%)", kpis["avg_purity"]),
        ("OEE (%)", kpis["oee"]),
        ("Horas Operativas", kpis["operating_hours"]),
        ("Costo Total ($)", kpis["total_cost"]),
        ("Ingreso Total ($)", kpis["total_revenue"]),
        ("Margen ($)", kpis["margin"]),
        ("Margen (%)", kpis["margin_pct"]),
        ("Lotes Completados", kpis["completed_batches"]),
        ("Lotes Totales", kpis["total_batches"]),
        ("Temp. Promedio", kpis["avg_temperature"]),
        ("Presión Promedio", kpis["avg_pressure"]),
    ]
    for idx, (label, val) in enumerate(kpi_rows, start=5):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = kpi_label_font
        ws[f"B{idx}"] = val if val is not None else "—"
        ws[f"A{idx}"].border = thin_border
        ws[f"B{idx}"].border = thin_border

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # ── Sheet 2: Batches
    ws2 = wb.create_sheet("Lotes")
    batch_headers = [
        "Lote", "Estado", "Inicio", "Fin", "Alimentación (kg)",
        "Producto (kg)", "Rendimiento (%)", "Pureza (%)", "Calidad",
        "Temp. Promedio", "Presión Promedio", "Costo ($)", "Ingreso ($)",
    ]
    for col, h in enumerate(batch_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, b in enumerate(batches, start=2):
        vals = [
            b["batch_number"], b["status"], b["actual_start"], b["actual_end"],
            b["feed_kg"], b["product_kg"], b["yield"], b["purity"],
            b["quality_grade"], b["avg_temp"], b["avg_press"],
            b["cost"], b["revenue"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col, value=v if v is not None else "—")
            cell.border = thin_border

    for col in range(1, len(batch_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 3: Alarm Statistics
    ws3 = wb.create_sheet("Alarmas")
    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Estadísticas de Alarmas"
    ws3["A1"].font = title_font

    alarm_rows = [
        ("Total Alarmas", alarm_stats["total"]),
        ("Críticas", alarm_stats["critical"]),
        ("Altas", alarm_stats["high"]),
        ("Tiempo Resp. Promedio (s)", alarm_stats["avg_response_time_s"]),
        ("% Reconocidas", alarm_stats["acknowledged_pct"]),
    ]
    for idx, (label, val) in enumerate(alarm_rows, start=3):
        ws3[f"A{idx}"] = label
        ws3[f"A{idx}"].font = kpi_label_font
        ws3[f"B{idx}"] = val if val is not None else "—"
        ws3[f"A{idx}"].border = thin_border
        ws3[f"B{idx}"].border = thin_border

    # Top alarm instruments
    ws3[f"A{len(alarm_rows) + 5}"] = "Top Instrumentos"
    ws3[f"A{len(alarm_rows) + 5}"].font = title_font
    start_row = len(alarm_rows) + 6
    for col, h in enumerate(["Instrumento", "Cantidad"], 1):
        cell = ws3.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for i, item in enumerate(alarm_stats["top_instruments"], start=1):
        ws3.cell(row=start_row + i, column=1, value=item["tag"]).border = thin_border
        ws3.cell(row=start_row + i, column=2, value=item["count"]).border = thin_border

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    # ── Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"PetroVision_Report_{report_type}_{start.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── 5. Dashboard summary (quick KPIs) ───────────────────────

@router.get("/summary")
def report_summary(
    process_id: int = Query(1),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Quick summary KPIs for last 24h, 7d, 30d."""
    now = datetime.now(timezone.utc)
    periods = {
        "last_24h": now - timedelta(hours=24),
        "last_7d": now - timedelta(days=7),
        "last_30d": now - timedelta(days=30),
    }

    result = {}
    for label, since in periods.items():
        result[label] = _compute_kpis(db, process_id, since, now)

    return result
